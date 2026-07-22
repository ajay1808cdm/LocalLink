import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

def generate_report():
    with open("test_cases.json", "r", encoding="utf-8") as f:
        data = json.load(f)

    df_all = pd.DataFrame(data)

    output_dir = "../Vulnerability Test Results"
    os.makedirs(output_dir, exist_ok=True)
    report_path = os.path.join(output_dir, "LocalLink_400_Vulnerability_Test_Report.xlsx")

    # Filter dataframes for different sheets
    df_critical = df_all[df_all["Severity"] == "Critical"]
    df_high = df_all[df_all["Severity"] == "High"]
    df_medium = df_all[df_all["Severity"] == "Medium"]
    df_low = df_all[df_all["Severity"] == "Low"]
    df_failed = df_all[df_all["Status"] == "Fail"]
    df_remediation = df_failed[["Test Case ID", "Category", "Root Cause", "Recommended Fix", "Severity", "Fix Applied"]]

    # We will write these using pandas ExcelWriter
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        # Create empty placeholder sheets first for the summary sheets
        pd.DataFrame([{"Executive Summary": "See formatting step"}]).to_excel(writer, sheet_name="Executive Summary", index=False)
        
        df_all.to_excel(writer, sheet_name="All 400 Test Cases", index=False)
        df_critical.to_excel(writer, sheet_name="Critical Findings", index=False)
        df_high.to_excel(writer, sheet_name="High Findings", index=False)
        df_medium.to_excel(writer, sheet_name="Medium Findings", index=False)
        df_low.to_excel(writer, sheet_name="Low Findings", index=False)
        df_failed.to_excel(writer, sheet_name="Failed Tests", index=False)
        df_remediation.to_excel(writer, sheet_name="Remediation Plan", index=False)
        
        pd.DataFrame([{"Tool": "Scan results summary will be here"}]).to_excel(writer, sheet_name="Tool Scan Results", index=False)
        pd.DataFrame([{"GitHub Actions Summary": "Run statistics"}]).to_excel(writer, sheet_name="GitHub Actions Summary", index=False)

    # Now load with openpyxl to apply formatting and formulas
    wb = load_workbook(report_path)

    # Styling definitions
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
    blocked_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Orange
    not_run_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Grey

    crit_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    high_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    med_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    low_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    info_fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

    def apply_formatting(ws, apply_colors=True):
        ws.freeze_panes = "A2"
        # Enable filters for all columns
        ws.auto_filter.ref = ws.dimensions
        
        # Set column widths and wrap text
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) # Cap width to 50
            ws.column_dimensions[column].width = adjusted_width

        if not apply_colors:
            return

        # Find Status and Severity columns
        status_col = None
        severity_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == "Status":
                status_col = idx + 1
            if cell.value == "Severity":
                severity_col = idx + 1

        for row in range(2, ws.max_row + 1):
            if status_col:
                cell = ws.cell(row=row, column=status_col)
                if cell.value == "Pass":
                    cell.fill = pass_fill
                elif cell.value == "Fail":
                    cell.fill = fail_fill
                elif cell.value == "Blocked":
                    cell.fill = blocked_fill
                elif cell.value == "Not Run":
                    cell.fill = not_run_fill
            
            if severity_col:
                cell = ws.cell(row=row, column=severity_col)
                if cell.value == "Critical":
                    cell.fill = crit_fill
                elif cell.value == "High":
                    cell.fill = high_fill
                elif cell.value == "Medium":
                    cell.fill = med_fill
                elif cell.value == "Low":
                    cell.fill = low_fill
                elif cell.value == "Informational":
                    cell.fill = info_fill

    # Apply standard formatting to data sheets
    for sheet_name in wb.sheetnames:
        if sheet_name not in ["Executive Summary", "Tool Scan Results", "GitHub Actions Summary"]:
            apply_formatting(wb[sheet_name])

    # Executive Summary Formulas
    ws_exec = wb["Executive Summary"]
    ws_exec.delete_rows(1, ws_exec.max_row)
    
    ws_exec["A1"] = "Executive Summary"
    ws_exec["A1"].font = Font(size=14, bold=True)
    
    # Use EXACT formulas pointing to 'All 400 Test Cases'
    ws_exec["A3"] = "Total tests"
    ws_exec["B3"] = "=COUNTA('All 400 Test Cases'!A2:A401)"
    
    ws_exec["A4"] = "Passed"
    ws_exec["B4"] = "=COUNTIF('All 400 Test Cases'!J2:J401, \"Pass\")"
    
    ws_exec["A5"] = "Failed"
    ws_exec["B5"] = "=COUNTIF('All 400 Test Cases'!J2:J401, \"Fail\")"
    
    ws_exec["A6"] = "Blocked"
    ws_exec["B6"] = "=COUNTIF('All 400 Test Cases'!J2:J401, \"Blocked\")"
    
    ws_exec["A7"] = "Not Run"
    ws_exec["B7"] = "=COUNTIF('All 400 Test Cases'!J2:J401, \"Not Run\")"
    
    ws_exec["A8"] = "Pass percentage"
    ws_exec["B8"] = "=IF(B3>0, B4/B3, 0)"
    ws_exec["B8"].number_format = '0.00%'

    ws_exec["A10"] = "Findings by Severity"
    ws_exec["A10"].font = Font(bold=True)
    
    ws_exec["A11"] = "Critical"
    ws_exec["B11"] = "=COUNTIFS('All 400 Test Cases'!K2:K401, \"Critical\", 'All 400 Test Cases'!J2:J401, \"Fail\")"
    ws_exec["A12"] = "High"
    ws_exec["B12"] = "=COUNTIFS('All 400 Test Cases'!K2:K401, \"High\", 'All 400 Test Cases'!J2:J401, \"Fail\")"
    ws_exec["A13"] = "Medium"
    ws_exec["B13"] = "=COUNTIFS('All 400 Test Cases'!K2:K401, \"Medium\", 'All 400 Test Cases'!J2:J401, \"Fail\")"
    ws_exec["A14"] = "Low"
    ws_exec["B14"] = "=COUNTIFS('All 400 Test Cases'!K2:K401, \"Low\", 'All 400 Test Cases'!J2:J401, \"Fail\")"

    ws_exec.column_dimensions["A"].width = 25
    ws_exec.column_dimensions["B"].width = 15

    wb.save(report_path)
    print(f"Report generated successfully at: {report_path}")

if __name__ == "__main__":
    generate_report()
